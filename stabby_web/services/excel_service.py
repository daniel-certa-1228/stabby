from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from contextlib import suppress


class ExcelService:

    @classmethod
    def generate_knife_excel(cls, knife_queryset):
        wb, ws = cls._create_workbook("Knife List")

        headers = [
            "Brand",
            "Knife",
            "Type",
            "# of Blades",
            "Blade Material",
            "Handle Material",
            "Lock",
            "Deployment",
            "Country",
            "Vendor",
            "Has Pocket Clip",
            "Purchased New",
            "Needs Work",
            "Date Entered",
        ]

        cls._append_headers(ws, headers)

        for row_idx, knife in enumerate(knife_queryset, start=2):
            col = 1
            ws.cell(row=row_idx, column=col, value=knife.brand)
            col += 1
            ws.cell(row=row_idx, column=col, value=knife.knife)
            col += 1
            ws.cell(row=row_idx, column=col, value=knife.knife_type)
            col += 1
            ws.cell(row=row_idx, column=col, value=knife.num_of_blades)
            col += 1
            ws.cell(row=row_idx, column=col, value=knife.blade_material)
            col += 1
            ws.cell(row=row_idx, column=col, value=knife.handle_material)
            col += 1
            ws.cell(row=row_idx, column=col, value=knife.lock_type)
            col += 1
            ws.cell(row=row_idx, column=col, value=knife.deployment_type)
            col += 1
            ws.cell(row=row_idx, column=col, value=knife.country)
            col += 1
            ws.cell(row=row_idx, column=col, value=knife.vendor)
            col += 1
            cls._write_boolean_cell(
                ws, row_idx, col, knife.has_pocket_clip, "Yes", "No"
            )
            col += 1
            cls._write_boolean_cell(
                ws, row_idx, col, knife.purchased_new, "New", "Used"
            )
            col += 1
            cls._write_boolean_cell(ws, row_idx, col, knife.needs_work)
            col += 1
            # Date col
            date_cell = ws.cell(
                row=row_idx, column=col, value=knife.create_date.replace(tzinfo=None)
            )
            date_cell.number_format = "m/d/yyyy"
            col += 1

        cls._auto_width_hack(ws, len(headers))

        # Manual widths for problem columns
        column_map = {
            "Deployment": 13,
            "Purchased New": 15,
            "Needs Work": 13,
            "Date Entered": 15,
        }

        cls._manual_width_set(ws, headers, column_map)

        ws.freeze_panes = "A2"

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()

    @classmethod
    def generate_sharpener_excel(cls, sharpener_queryset):
        wb, ws = cls._create_workbook("Sharpener List")

        headers = [
            "Brand",
            "Sharpener",
            "Cutting Agent",
            "Bonding Agent",
            "Length",
            "Width",
            "Country",
            "Friable",
        ]

        cls._append_headers(ws, headers)

        for row_idx, sharpener in enumerate(sharpener_queryset, start=2):
            col = 1
            ws.cell(row=row_idx, column=col, value=sharpener.brand)
            col += 1
            ws.cell(row=row_idx, column=col, value=sharpener.sharpener)
            col += 1
            ws.cell(row=row_idx, column=col, value=sharpener.cutting_agent)
            col += 1
            ws.cell(row=row_idx, column=col, value=sharpener.bonding_agent)
            col += 1
            ws.cell(row=row_idx, column=col, value=sharpener.length)
            col += 1
            ws.cell(row=row_idx, column=col, value=sharpener.width)
            col += 1
            ws.cell(row=row_idx, column=col, value=sharpener.country)
            col += 1
            cls._write_boolean_cell(ws, row_idx, col, sharpener.is_friable)
            col += 1

        cls._auto_width_hack(ws, len(headers))

        column_map = {
            "Country": 9,
        }

        cls._manual_width_set(ws, headers, column_map)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()

    @classmethod
    def _create_workbook(cls, title: str):
        wb = Workbook()
        ws = wb.active
        ws.title = title
        return wb, ws

    @classmethod
    def _append_headers(cls, ws, headers: list[str]):
        header_font = Font(bold=True, color="FFFFFF")

        header_fill = PatternFill(
            fill_type="solid", start_color="12387A", end_color="12387A"
        )

        ws.append(headers)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill

    @classmethod
    def _write_boolean_cell(
        cls, ws, row: int, col: int, value: bool, true_display="Yes", false_display="No"
    ):
        cell_value = true_display if value else false_display

        cell = ws.cell(row=row, column=col, value=cell_value)

        if not value:
            cell.font = Font(color="A0A0A0")

    @classmethod
    def _auto_width_hack(cls, ws, num_columns: int):
        for col_idx in range(1, num_columns + 1):
            col_letter = get_column_letter(col_idx)

            max_length = 0

            for row in ws.iter_rows(
                min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx
            ):
                for cell in row:
                    if cell.value:
                        with suppress(AttributeError, TypeError):
                            # Strip the value, convert to string, and measure its length
                            max_length = max(max_length, len(str(cell.value).strip()))

            ws.column_dimensions[col_letter].width = max_length

    @classmethod
    def _manual_width_set(cls, ws, headers: list[str], column_map: dict[str, int]):
        for header_name, width in column_map.items():
            col_letter = get_column_letter(headers.index(header_name) + 1)
            ws.column_dimensions[col_letter].width = width
